# SYSCOHADA Professional Dashboard Generator

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Function to create individual sheets

def create_sheet(workbook, sheet_name, data, title):
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(title)
    for row in data:
        sheet.append(row)

    # Apply professional styling
    for cell in sheet['1:1']:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Main Function to generate dashboard

def generate_dashboard():
    # Create a workbook and add title sheet
    workbook = openpyxl.Workbook()
    title_sheet = workbook.active
    title_sheet.title = "Dashboard"
    title_sheet['A1'] = "SYSCOHADA Accounting Dashboard"
    title_sheet['A1'].font = Font(size=14, bold=True)

    # Sample data for various sheets
    data_overview = [["Metric", "Value"], ["Total Revenue", 100000], ["Total Expenses", 60000], ["Net Profit", 40000]]
    data_income_statement = [["Item", "Amount"], ["Sales Revenue", 150000], ["Cost of Goods Sold", 50000], ["Gross Profit", 100000]]

    # Add sheets with data
    create_sheet(workbook, "Overview", data_overview, ["Overview Metrics"])
    create_sheet(workbook, "Income Statement", data_income_statement, ["Income Statement"])

    # Save the workbook
    workbook.save("SYSCOHADA_Dashboard.xlsx")

# Execute the generator
if __name__ == '__main__':
    generate_dashboard()