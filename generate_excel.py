"""
Génère un fichier Excel (.xlsx) du Plan Comptable SYSCOHADA Révisé (2017)
avec mise en forme, onglets par classe et récapitulatif.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Données : Plan Comptable SYSCOHADA Révisé ────────────────────────────────

ACCOUNTS = [
    # (code, intitulé, classe, type, nature)
    # CLASSE 1
    ("101","Capital social","1","Passif","Créditeur"),
    ("102","Capital par dotation","1","Passif","Créditeur"),
    ("104","Comptes d'apporteurs – opérations sur capital","1","Passif","Créditeur"),
    ("105","Primes liées au capital social","1","Passif","Créditeur"),
    ("106","Écarts de réévaluation","1","Passif","Créditeur"),
    ("111","Réserve légale","1","Passif","Créditeur"),
    ("112","Réserves statutaires ou contractuelles","1","Passif","Créditeur"),
    ("118","Autres réserves","1","Passif","Créditeur"),
    ("120","Report à nouveau (solde créditeur)","1","Passif","Créditeur"),
    ("129","Report à nouveau (solde débiteur)","1","Passif","Débiteur"),
    ("130","Résultat net de l'exercice – bénéfice","1","Passif","Créditeur"),
    ("139","Résultat net de l'exercice – perte","1","Passif","Débiteur"),
    ("141","Subventions d'équipement","1","Passif","Créditeur"),
    ("161","Emprunts obligataires","1","Passif","Créditeur"),
    ("162","Emprunts auprès des établissements de crédit","1","Passif","Créditeur"),
    ("163","Avances reçues de l'État et des organismes publics","1","Passif","Créditeur"),
    ("164","Emprunts auprès des associés","1","Passif","Créditeur"),
    ("165","Dépôts et cautionnements reçus","1","Passif","Créditeur"),
    ("181","Dettes de crédit-bail","1","Passif","Créditeur"),
    ("191","Provisions pour risques","1","Passif","Créditeur"),
    ("194","Provisions pour charges","1","Passif","Créditeur"),
    # CLASSE 2
    ("201","Frais d'établissement","2","Actif","Débiteur"),
    ("202","Frais de recherche et développement","2","Actif","Débiteur"),
    ("203","Logiciels","2","Actif","Débiteur"),
    ("204","Brevets, licences, marques","2","Actif","Débiteur"),
    ("205","Fonds commercial","2","Actif","Débiteur"),
    ("211","Terrains d'exploitation","2","Actif","Débiteur"),
    ("213","Bâtiments sur sol propre","2","Actif","Débiteur"),
    ("214","Constructions sur sol d'autrui","2","Actif","Débiteur"),
    ("215","Installations et agencements","2","Actif","Débiteur"),
    ("221","Matériel et outillage industriel","2","Actif","Débiteur"),
    ("222","Matériel d'emballage récupérable","2","Actif","Débiteur"),
    ("224","Matériel et mobilier de bureau","2","Actif","Débiteur"),
    ("225","Matériel informatique","2","Actif","Débiteur"),
    ("226","Matériel de transport","2","Actif","Débiteur"),
    ("228","Autres matériels","2","Actif","Débiteur"),
    ("261","Titres de participation","2","Actif","Débiteur"),
    ("271","Prêts et créances rattachés à des participations","2","Actif","Débiteur"),
    ("272","Prêts au personnel","2","Actif","Débiteur"),
    ("274","Dépôts et cautionnements versés","2","Actif","Débiteur"),
    ("281","Amortissements – immobilisations incorporelles","2","Actif","Créditeur"),
    ("283","Amortissements – bâtiments","2","Actif","Créditeur"),
    ("284","Amortissements – matériels","2","Actif","Créditeur"),
    # CLASSE 3
    ("31","Marchandises","3","Actif","Débiteur"),
    ("321","Matières premières","3","Actif","Débiteur"),
    ("322","Matières et fournitures consommables","3","Actif","Débiteur"),
    ("33","Autres approvisionnements","3","Actif","Débiteur"),
    ("34","Produits en cours","3","Actif","Débiteur"),
    ("35","Services en cours","3","Actif","Débiteur"),
    ("36","Produits finis","3","Actif","Débiteur"),
    ("37","Produits intermédiaires et résiduels","3","Actif","Débiteur"),
    # CLASSE 4
    ("401","Fournisseurs","4","Passif","Créditeur"),
    ("402","Fournisseurs – effets à payer","4","Passif","Créditeur"),
    ("408","Fournisseurs – factures non reçues","4","Passif","Créditeur"),
    ("409","Fournisseurs – avances et acomptes versés","4","Actif","Débiteur"),
    ("411","Clients","4","Actif","Débiteur"),
    ("412","Clients – effets à recevoir","4","Actif","Débiteur"),
    ("418","Clients – produits à recevoir","4","Actif","Débiteur"),
    ("419","Clients – avances et acomptes reçus","4","Passif","Créditeur"),
    ("421","Personnel – avances et acomptes","4","Actif","Débiteur"),
    ("422","Personnel – rémunérations dues","4","Passif","Créditeur"),
    ("431","Organismes sociaux","4","Passif","Créditeur"),
    ("441","État – impôt sur les bénéfices","4","Passif","Créditeur"),
    ("442","État – autres impôts et taxes","4","Passif","Créditeur"),
    ("443","État – TVA facturée","4","Passif","Créditeur"),
    ("445","État – TVA récupérable","4","Actif","Débiteur"),
    ("446","État – crédit de TVA","4","Actif","Débiteur"),
    ("471","Débiteurs divers","4","Actif","Débiteur"),
    ("472","Créditeurs divers","4","Passif","Créditeur"),
    ("476","Charges constatées d'avance","4","Actif","Débiteur"),
    ("477","Produits constatés d'avance","4","Passif","Créditeur"),
    ("481","Fournisseurs d'investissement","4","Passif","Créditeur"),
    ("485","Créances sur cessions d'immobilisations","4","Actif","Débiteur"),
    # CLASSE 5
    ("511","Valeurs à l'encaissement","5","Actif","Débiteur"),
    ("512","Chèques à encaisser","5","Actif","Débiteur"),
    ("521","Banque locale 1","5","Actif","Débiteur"),
    ("522","Banque locale 2","5","Actif","Débiteur"),
    ("531","Chèques postaux","5","Actif","Débiteur"),
    ("571","Caisse siège social","5","Actif","Débiteur"),
    ("572","Caisse annexe","5","Actif","Débiteur"),
    ("581","Virements internes","5","Actif","Débiteur"),
    # CLASSE 6
    ("601","Achats de marchandises","6","Charge","Débiteur"),
    ("602","Achats de matières premières","6","Charge","Débiteur"),
    ("604","Achats stockés – matières et fournitures consommables","6","Charge","Débiteur"),
    ("605","Autres achats","6","Charge","Débiteur"),
    ("608","Achats d'emballages","6","Charge","Débiteur"),
    ("609","Rabais, remises et ristournes obtenus sur achats","6","Charge","Créditeur"),
    ("611","Transports sur achats","6","Charge","Débiteur"),
    ("612","Transports sur ventes","6","Charge","Débiteur"),
    ("613","Transports pour le compte de tiers","6","Charge","Débiteur"),
    ("614","Transports du personnel","6","Charge","Débiteur"),
    ("621","Sous-traitance générale","6","Charge","Débiteur"),
    ("622","Locations et charges locatives","6","Charge","Débiteur"),
    ("623","Redevances de crédit-bail","6","Charge","Débiteur"),
    ("624","Entretien, réparations et maintenance","6","Charge","Débiteur"),
    ("625","Primes d'assurance","6","Charge","Débiteur"),
    ("626","Études, recherches et documentation","6","Charge","Débiteur"),
    ("627","Publicité et relations publiques","6","Charge","Débiteur"),
    ("628","Frais de télécommunications","6","Charge","Débiteur"),
    ("629","Autres services extérieurs A","6","Charge","Débiteur"),
    ("631","Frais bancaires","6","Charge","Débiteur"),
    ("632","Rémunérations d'intermédiaires","6","Charge","Débiteur"),
    ("633","Frais de formation du personnel","6","Charge","Débiteur"),
    ("634","Cadeaux à la clientèle","6","Charge","Débiteur"),
    ("635","Frais de voyage et déplacements","6","Charge","Débiteur"),
    ("637","Frais de réceptions","6","Charge","Débiteur"),
    ("641","Impôts et taxes directs","6","Charge","Débiteur"),
    ("642","Droits d'enregistrement","6","Charge","Débiteur"),
    ("645","Taxes sur chiffre d'affaires non récupérables","6","Charge","Débiteur"),
    ("651","Pertes sur créances irrécouvrables","6","Charge","Débiteur"),
    ("658","Charges diverses","6","Charge","Débiteur"),
    ("661","Rémunérations directes versées au personnel","6","Charge","Débiteur"),
    ("662","Rémunérations des congés","6","Charge","Débiteur"),
    ("663","Indemnités et avantages divers","6","Charge","Débiteur"),
    ("664","Charges sociales","6","Charge","Débiteur"),
    ("671","Intérêts des emprunts","6","Charge","Débiteur"),
    ("672","Intérêts des dettes de crédit-bail","6","Charge","Débiteur"),
    ("673","Escomptes accordés","6","Charge","Débiteur"),
    ("674","Pertes de change","6","Charge","Débiteur"),
    ("681","Dotations aux amortissements des immobilisations","6","Charge","Débiteur"),
    ("691","Dotations aux provisions pour risques à court terme","6","Charge","Débiteur"),
    # CLASSE 7
    ("701","Ventes de marchandises","7","Produit","Créditeur"),
    ("702","Ventes de produits finis","7","Produit","Créditeur"),
    ("703","Ventes de produits intermédiaires","7","Produit","Créditeur"),
    ("704","Ventes de produits résiduels","7","Produit","Créditeur"),
    ("705","Travaux facturés","7","Produit","Créditeur"),
    ("706","Services vendus","7","Produit","Créditeur"),
    ("707","Produits accessoires","7","Produit","Créditeur"),
    ("709","Rabais, remises et ristournes accordés","7","Produit","Débiteur"),
    ("711","Subventions d'exploitation","7","Produit","Créditeur"),
    ("721","Immobilisations produites par l'entreprise","7","Produit","Créditeur"),
    ("731","Variations de stocks de biens produits","7","Produit","Créditeur"),
    ("741","Revenus des titres de participation","7","Produit","Créditeur"),
    ("742","Revenus des autres immobilisations financières","7","Produit","Créditeur"),
    ("751","Récupérations sur créances amorties","7","Produit","Créditeur"),
    ("752","Produits sur exercices antérieurs","7","Produit","Créditeur"),
    ("758","Produits divers","7","Produit","Créditeur"),
    ("771","Intérêts de prêts","7","Produit","Créditeur"),
    ("772","Revenus de placements","7","Produit","Créditeur"),
    ("773","Escomptes obtenus","7","Produit","Créditeur"),
    ("774","Gains de change","7","Produit","Créditeur"),
    ("781","Transferts de charges d'exploitation","7","Produit","Créditeur"),
    ("791","Reprises de provisions pour risques","7","Produit","Créditeur"),
    # CLASSE 8
    ("811","Valeurs comptables des cessions d'immobilisations","8","Charge HAO","Débiteur"),
    ("821","Produits des cessions d'immobilisations","8","Produit HAO","Créditeur"),
    ("831","Charges HAO sur exercices antérieurs","8","Charge HAO","Débiteur"),
    ("832","Pénalités et amendes fiscales et pénales","8","Charge HAO","Débiteur"),
    ("838","Autres charges HAO","8","Charge HAO","Débiteur"),
    ("841","Produits HAO sur exercices antérieurs","8","Produit HAO","Créditeur"),
    ("848","Autres produits HAO","8","Produit HAO","Créditeur"),
    ("871","Participation des travailleurs","8","Charge HAO","Débiteur"),
    ("881","Impôt sur le résultat","8","Charge HAO","Débiteur"),
]

CLASS_INFO = {
    "1": ("Ressources Durables",          "1A3A6B", "FFFFFF"),
    "2": ("Actif Immobilisé",             "1565C0", "FFFFFF"),
    "3": ("Stocks",                        "E65100", "FFFFFF"),
    "4": ("Comptes de Tiers",             "F57F17", "FFFFFF"),
    "5": ("Trésorerie",                   "2E7D32", "FFFFFF"),
    "6": ("Charges des Act. Ordinaires",  "B71C1C", "FFFFFF"),
    "7": ("Produits des Act. Ordinaires", "1B5E20", "FFFFFF"),
    "8": ("Charges & Produits HAO",       "4A148C", "FFFFFF"),
}

TYPE_COLORS = {
    "Actif":       "BBDEFB",
    "Passif":      "C8E6C9",
    "Charge":      "FFCDD2",
    "Produit":     "DCEDC8",
    "Charge HAO":  "E1BEE7",
    "Produit HAO": "D1C4E9",
}

# ─── Styles helpers ───────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_row(ws, row, headers, bg, fg="FFFFFF"):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=10)
        c.alignment = center()
        c.border = thin_border()

# ─── Workbook ─────────────────────────────────────────────────────────────────

wb = Workbook()

# ─── Onglet : Plan Complet ────────────────────────────────────────────────────

ws_all = wb.active
ws_all.title = "Plan Complet"
ws_all.sheet_view.showGridLines = False

# Title
ws_all.merge_cells("A1:E1")
title_cell = ws_all["A1"]
title_cell.value = "PLAN COMPTABLE GÉNÉRAL – SYSCOHADA RÉVISÉ 2017"
title_cell.fill = fill("1A3A6B")
title_cell.font = font(bold=True, color="FFFFFF", size=14)
title_cell.alignment = center()
ws_all.row_dimensions[1].height = 32

ws_all.merge_cells("A2:E2")
sub_cell = ws_all["A2"]
sub_cell.value = "Organisation pour l'Harmonisation en Afrique du Droit des Affaires (OHADA)"
sub_cell.fill = fill("D4A017")
sub_cell.font = font(bold=True, color="FFFFFF", size=10)
sub_cell.alignment = center()
ws_all.row_dimensions[2].height = 20

# Column headers
headers = ["Code", "Intitulé du Compte", "Classe", "Type", "Nature"]
header_row(ws_all, 3, headers, "263238", "FFFFFF")
ws_all.row_dimensions[3].height = 22

# Data
current_class = None
row = 4
for code, label, cls, typ, nature in ACCOUNTS:
    if cls != current_class:
        # Class separator
        ws_all.merge_cells(f"A{row}:E{row}")
        cl_name, bg, fg = CLASS_INFO[cls]
        c = ws_all.cell(row=row, column=1,
                        value=f"  CLASSE {cls}  –  {cl_name}")
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=10)
        c.alignment = left()
        ws_all.row_dimensions[row].height = 18
        row += 1
        current_class = cls

    type_bg = TYPE_COLORS.get(typ, "FFFFFF")
    row_data = [code, label, f"Classe {cls}", typ, nature]
    for col, val in enumerate(row_data, 1):
        c = ws_all.cell(row=row, column=col, value=val)
        c.fill = fill(type_bg)
        c.font = font(size=9)
        c.alignment = left() if col == 2 else center()
        c.border = thin_border()
    ws_all.row_dimensions[row].height = 15
    row += 1

# Column widths
set_col_width(ws_all, "A", 10)
set_col_width(ws_all, "B", 60)
set_col_width(ws_all, "C", 14)
set_col_width(ws_all, "D", 16)
set_col_width(ws_all, "E", 14)

# Freeze header
ws_all.freeze_panes = "A4"

# Auto-filter
ws_all.auto_filter.ref = f"A3:E{row-1}"

# ─── Onglets par Classe ───────────────────────────────────────────────────────

for cls, (cls_name, bg_hex, fg_hex) in CLASS_INFO.items():
    ws = wb.create_sheet(title=f"Cl.{cls} – {cls_name[:12]}")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"CLASSE {cls} – {cls_name.upper()}"
    t.fill = fill(bg_hex)
    t.font = font(bold=True, color=fg_hex, size=12)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = "SYSCOHADA Révisé 2017 – Plan Comptable Général OHADA"
    s.fill = fill("ECEFF1")
    s.font = font(color="546E7A", size=9)
    s.alignment = center()
    ws.row_dimensions[2].height = 16

    header_row(ws, 3, ["Code", "Intitulé du Compte", "Type", "Nature", "Solde Normal"], bg_hex, fg_hex)
    ws.row_dimensions[3].height = 20

    r = 4
    cls_accounts = [(c, l, t, n) for c, l, cl, t, n in ACCOUNTS if cl == cls]
    for i, (code, label, typ, nature) in enumerate(cls_accounts):
        type_bg = TYPE_COLORS.get(typ, "FFFFFF")
        alt = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        row_vals = [code, label, typ, nature, "Débit" if nature == "Débiteur" else "Crédit"]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(type_bg if col == 3 else alt)
            c.font = font(bold=(col == 1), size=9,
                          color="1565C0" if col == 1 else "000000")
            c.alignment = left() if col == 2 else center()
            c.border = thin_border()
        ws.row_dimensions[r].height = 15
        r += 1

    # Totals row
    ws.merge_cells(f"A{r}:B{r}")
    tot = ws.cell(row=r, column=1,
                  value=f"Total : {len(cls_accounts)} comptes – Classe {cls}")
    tot.fill = fill(bg_hex)
    tot.font = font(bold=True, color=fg_hex, size=9)
    tot.alignment = center()
    ws.row_dimensions[r].height = 16

    set_col_width(ws, "A", 10)
    set_col_width(ws, "B", 58)
    set_col_width(ws, "C", 16)
    set_col_width(ws, "D", 14)
    set_col_width(ws, "E", 14)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{r-1}"

# ─── Onglet : Ratios Financiers OHADA ────────────────────────────────────────

RATIOS = [
    # (catégorie, intitulé, formule, norme/cible, interprétation)
    ("Rentabilité","Rentabilité des CP (ROE)",
     "Résultat Net / Capitaux Propres × 100",
     "≥ 10 %","Rendement généré pour les actionnaires"),
    ("Rentabilité","Rentabilité des Actifs (ROA)",
     "Résultat Net / Total Actif × 100",
     "≥ 3 %","Efficacité d'utilisation de l'ensemble des actifs"),
    ("Rentabilité","Marge Nette",
     "Résultat Net / Chiffre d'Affaires × 100",
     "≥ 3 %","Part du CA transformée en bénéfice net"),
    ("Rentabilité","Marge Brute d'Exploitation",
     "(CA – Coût des Ventes) / CA × 100",
     "≥ 20 %","Marge dégagée avant frais généraux"),
    ("Rentabilité","Taux de Charges",
     "Total Charges / Total Produits × 100",
     "< 90 %","Part des produits absorbée par les charges"),

    ("Liquidité","Liquidité Générale (Current Ratio)",
     "Actif Circulant / Passif Circulant",
     "≥ 1,5x","Capacité à honorer les dettes à court terme"),
    ("Liquidité","Liquidité Réduite (Quick Ratio)",
     "(Actif Circulant – Stocks) / Passif Circulant",
     "≥ 1,0x","Liquidité sans les stocks (moins liquides)"),
    ("Liquidité","Liquidité Immédiate (Cash Ratio)",
     "Trésorerie / Passif Circulant",
     "≥ 0,2x","Capacité à payer immédiatement"),

    ("Solvabilité","Autonomie Financière",
     "Capitaux Propres / Total Passif × 100",
     "≥ 30 %","Part du financement assurée par les fonds propres"),
    ("Solvabilité","Ratio d'Endettement (Gearing)",
     "Dettes Totales / Capitaux Propres × 100",
     "< 150 %","Niveau d'endettement vs fonds propres"),
    ("Solvabilité","Capacité de Remboursement",
     "Dettes Financières / CAF",
     "< 5 ans","Années nécessaires pour rembourser via la CAF"),
    ("Solvabilité","Couverture des Immobilisations",
     "Ressources Stables / Emplois Stables × 100",
     "≥ 100 %","Les ressources stables financent les emplois stables"),

    ("Activité","Rotation des Stocks",
     "Chiffre d'Affaires / Stock",
     "4 à 12x/an","Fréquence de renouvellement des stocks"),
    ("Activité","Délai Recouvrement Clients",
     "(Créances Clients / CA) × 365",
     "≤ 60 jours","Délai moyen d'encaissement des créances"),
    ("Activité","Délai Paiement Fournisseurs",
     "(Dettes Fournisseurs / Achats) × 365",
     "30 à 90 jours","Délai moyen de paiement des fournisseurs"),
    ("Activité","Rotation de l'Actif Total",
     "Chiffre d'Affaires / Total Actif",
     "0,5 à 2x","Efficacité d'utilisation de l'actif total"),

    ("Équilibre OHADA","FRNG – Fonds de Roulement Net Global",
     "Ressources Stables – Emplois Stables",
     "≥ 0","Excédent des ressources stables sur emplois stables"),
    ("Équilibre OHADA","BFR – Besoin en Fonds de Roulement",
     "Actif Circulant (hors tréso) – Passif Circulant (hors tréso)",
     "BFR < FRNG","Besoin de financement du cycle d'exploitation"),
    ("Équilibre OHADA","Trésorerie Nette",
     "FRNG – BFR",
     "≥ 0","Position de trésorerie structurelle (équilibre financier)"),
]

RATIO_CAT_COLORS = {
    "Rentabilité":      ("1A237E","E8EAF6"),
    "Liquidité":        ("1B5E20","E8F5E9"),
    "Solvabilité":      ("B71C1C","FFEBEE"),
    "Activité":         ("E65100","FFF3E0"),
    "Équilibre OHADA":  ("4A148C","F3E5F5"),
}

ws_r = wb.create_sheet(title="Ratios Financiers OHADA")
ws_r.sheet_view.showGridLines = False

ws_r.merge_cells("A1:E1")
t = ws_r["A1"]
t.value = "RATIOS FINANCIERS – NORMES OHADA (SYSCOHADA RÉVISÉ)"
t.fill = fill("1A3A6B")
t.font = font(bold=True, color="FFFFFF", size=13)
t.alignment = center()
ws_r.row_dimensions[1].height = 30

header_row(ws_r, 2,
           ["Catégorie","Ratio","Formule","Norme / Cible","Interprétation"],
           "263238","FFFFFF")
ws_r.row_dimensions[2].height = 20

r = 3
current_cat = None
for cat, ratio, formula, norm, interp in RATIOS:
    fg_cat, bg_cat = RATIO_CAT_COLORS.get(cat, ("000000","FFFFFF"))
    if cat != current_cat:
        ws_r.merge_cells(f"A{r}:E{r}")
        c = ws_r.cell(row=r, column=1, value=f"  {cat.upper()}")
        c.fill = fill(fg_cat)
        c.font = font(bold=True, color="FFFFFF", size=10)
        c.alignment = left()
        ws_r.row_dimensions[r].height = 16
        r += 1
        current_cat = cat

    for col, val in enumerate([cat, ratio, formula, norm, interp], 1):
        c = ws_r.cell(row=r, column=col, value=val)
        c.fill = fill(bg_cat)
        c.font = font(size=9, color="000000")
        c.alignment = left() if col in (2,3,5) else center()
        c.border = thin_border()
    ws_r.row_dimensions[r].height = 30
    r += 1

set_col_width(ws_r, "A", 18)
set_col_width(ws_r, "B", 30)
set_col_width(ws_r, "C", 45)
set_col_width(ws_r, "D", 18)
set_col_width(ws_r, "E", 45)
ws_r.freeze_panes = "A3"

# ─── Onglet : Journal (modèle vierge) ────────────────────────────────────────

ws_j = wb.create_sheet(title="Journal (modèle)")
ws_j.sheet_view.showGridLines = False

ws_j.merge_cells("A1:H1")
t = ws_j["A1"]
t.value = "JOURNAL GÉNÉRAL – SYSCOHADA RÉVISÉ"
t.fill = fill("1A3A6B")
t.font = font(bold=True, color="FFFFFF", size=12)
t.alignment = center()
ws_j.row_dimensions[1].height = 26

header_row(ws_j, 2,
           ["Date","N° Pièce","Journal","N° Compte","Intitulé du Compte","Libellé de l'Opération","Débit","Crédit"],
           "37474F","FFFFFF")
ws_j.row_dimensions[2].height = 22

# Sample entry
samples = [
    ("01/01/2024","AN-001","AN","521","Banque locale 1","Bilan d'ouverture – trésorerie initiale","5 000 000",""),
    ("01/01/2024","AN-001","AN","101","Capital social","Bilan d'ouverture – capital","","5 000 000"),
    ("15/01/2024","AC-001","AC","601","Achats de marchandises","Achat stock initial fournisseur A","1 200 000",""),
    ("15/01/2024","AC-001","AC","401","Fournisseurs","Achat stock initial fournisseur A","","1 200 000"),
    ("31/01/2024","VT-001","VT","411","Clients","Vente client B – facture 001","800 000",""),
    ("31/01/2024","VT-001","VT","701","Ventes de marchandises","Vente client B – facture 001","","800 000"),
]

for i, row_data in enumerate(samples, 3):
    bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
    for col, val in enumerate(row_data, 1):
        c = ws_j.cell(row=i, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(size=9)
        c.alignment = center() if col in (1,2,3,4,7,8) else left()
        c.border = thin_border()

# Empty rows for input
for i in range(len(samples)+3, len(samples)+53):
    bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
    for col in range(1, 9):
        c = ws_j.cell(row=i, column=col, value="")
        c.fill = fill(bg)
        c.border = thin_border()
        c.font = font(size=9)

# Totals row
tot_row = len(samples) + 53
ws_j.merge_cells(f"A{tot_row}:F{tot_row}")
c = ws_j.cell(row=tot_row, column=1, value="TOTAUX")
c.fill = fill("263238")
c.font = font(bold=True, color="FFFFFF", size=10)
c.alignment = center()
for col in (7, 8):
    start = 3
    end = tot_row - 1
    c = ws_j.cell(row=tot_row, column=col,
                  value=f"=SUM({get_column_letter(col)}{start}:{get_column_letter(col)}{end})")
    c.fill = fill("263238")
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = center()
    c.border = thin_border()

set_col_width(ws_j, "A", 12)
set_col_width(ws_j, "B", 14)
set_col_width(ws_j, "C", 10)
set_col_width(ws_j, "D", 10)
set_col_width(ws_j, "E", 40)
set_col_width(ws_j, "F", 40)
set_col_width(ws_j, "G", 16)
set_col_width(ws_j, "H", 16)
ws_j.freeze_panes = "A3"

# ─── Sauvegarde ───────────────────────────────────────────────────────────────

output_path = "/home/user/Tableau-de-bord-/Plan_Comptable_SYSCOHADA_Revise.xlsx"
wb.save(output_path)
print(f"Fichier généré : {output_path}")
print(f"Onglets : {[ws.title for ws in wb.worksheets]}")
print(f"Total comptes : {len(ACCOUNTS)}")
print(f"Total ratios  : {len(RATIOS)}")
